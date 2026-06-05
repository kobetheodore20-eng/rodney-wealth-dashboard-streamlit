from __future__ import annotations

import sys
from pathlib import Path
import re

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from app.config import DATA_DIR, SOURCE_WORKBOOK  # noqa: E402


SHEETS = {
    "executive_dashboard": ("01 Executive Dashboard", "As of"),
    "balance_sheet": ("02 Balance Sheet", "Asset class"),
    "entity_ownership": ("02A Entity Ownership View", "Asset / bucket"),
    "liquidity_buckets": ("03 Liquidity Buckets", "Bucket"),
    "property_register": ("04 Property Register", "Property"),
    "loan_offset_register": ("05 Loan Offset Register", "Property"),
    "property_pnl": ("06 Property P&L", "Property"),
    "keith_statements": ("06A Keith Statements", "Statement #"),
    "keith_maintenance": ("06B Keith Maintenance", "Record type"),
    "keith_coverage": ("06C Keith Coverage", "Control item"),
    "land_tax_control": ("06D Land Tax Control", "Property / entity"),
    "last_month_cashflow": ("06E Last Month Cashflow", "Property"),
    "investment_register": ("07 Investment Register", "Asset / sleeve"),
    "listed_share_snapshot": ("07A Listed Share Snapshot", "Ticker"),
    "share_transactions": ("07B Share Transactions", "Date"),
    "fy_share_returns": ("07C FY Share Returns", "Australian FY"),
    "share_growth": ("07D Consolidated Share Growth", "Australian FY"),
    "benchmark_policy": ("07E Benchmark Policy", "Lens"),
    "super_retirement": ("08 Super Retirement", "Item"),
    "risk_dashboard": ("09 Risk Dashboard", "Risk metric"),
    "stress_tests": ("10 Stress Tests", "Scenario"),
    "decision_log": ("11 Decision Log", "Decision ID"),
    "evidence_register": ("12 Evidence Register", "Evidence ID"),
    "monthly_tracking": ("13 Monthly Tracking", "Date"),
}


def slugify(value: str) -> str:
    value = re.sub(r"[^a-zA-Z0-9]+", "_", value.lower()).strip("_")
    return value or "sheet"


def clean_frame(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.dropna(how="all")
    frame = frame.loc[:, ~frame.columns.astype(str).str.startswith("Unnamed")]
    frame = frame.dropna(how="all")
    frame.columns = [str(col).strip() for col in frame.columns]
    return frame.reset_index(drop=True)


def read_raw_sheet(sheet_name: str) -> pd.DataFrame:
    raw = pd.read_excel(SOURCE_WORKBOOK, sheet_name=sheet_name, header=None)
    raw = raw.dropna(how="all")
    raw = raw.loc[:, ~raw.isna().all(axis=0)]
    raw = raw.fillna("")
    if raw.empty:
        return raw
    raw.columns = [f"Column {idx + 1}" for idx in range(len(raw.columns))]
    return raw.reset_index(drop=True)


def read_sheet_table(sheet_name: str, header_marker: str) -> pd.DataFrame:
    raw = pd.read_excel(SOURCE_WORKBOOK, sheet_name=sheet_name, header=None)
    header_row = None
    for idx, row in raw.iterrows():
        values = [str(value).strip() for value in row.tolist() if pd.notna(value)]
        if header_marker in values:
            header_row = idx
            break

    if header_row is None:
        print(f"Warning: could not find header marker {header_marker!r} in {sheet_name}")
        return pd.DataFrame()

    headers = raw.iloc[header_row].tolist()
    frame = raw.iloc[header_row + 1 :].copy()
    frame.columns = headers
    return clean_frame(frame)


def formula_audit() -> dict[str, dict[str, object]]:
    formula_workbook = load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=False)
    value_workbook = load_workbook(SOURCE_WORKBOOK, read_only=True, data_only=True)
    audit: dict[str, dict[str, object]] = {}
    for worksheet in formula_workbook.worksheets:
        value_sheet = value_workbook[worksheet.title]
        formula_cells = 0
        cached_formula_blanks = 0
        populated_cells = 0
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    populated_cells += 1
                if cell.data_type == "f":
                    formula_cells += 1
                    cached_value = value_sheet[cell.coordinate].value
                    if cached_value in (None, ""):
                        cached_formula_blanks += 1
        if not formula_cells:
            fidelity = "No formulas detected"
        elif cached_formula_blanks:
            fidelity = "Formula cache incomplete - open/recalculate in Excel before import"
        else:
            fidelity = "Decision-grade cached formula values"
        audit[worksheet.title] = {
            "Formula cells": formula_cells,
            "Formula cached blanks": cached_formula_blanks,
            "Populated cells": populated_cells,
            "Formula fidelity": fidelity,
        }
    formula_workbook.close()
    value_workbook.close()
    return audit


def enforce_formula_fidelity(audit: dict[str, dict[str, object]]) -> None:
    operating_sheets = {sheet_name for sheet_name, _header_marker in SHEETS.values()}
    incomplete = [
        sheet
        for sheet in sorted(operating_sheets)
        if int(audit.get(sheet, {}).get("Formula cached blanks", 0) or 0) > 0
    ]
    if incomplete:
        joined = ", ".join(incomplete)
        raise ValueError(f"Workbook formula cache incomplete for operating sheets: {joined}. Open/recalculate in Excel before import.")


def main() -> None:
    if not SOURCE_WORKBOOK.exists():
        raise FileNotFoundError(f"Workbook not found: {SOURCE_WORKBOOK}")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    sheet_index = []
    workbook = pd.ExcelFile(SOURCE_WORKBOOK)
    audit = formula_audit()
    enforce_formula_fidelity(audit)
    for sheet_name in workbook.sheet_names:
        output_name = f"workbook_{slugify(sheet_name)}"
        frame = read_raw_sheet(sheet_name)
        frame.to_csv(DATA_DIR / f"{output_name}.csv", index=False)
        sheet_index.append(
            {
                "Sheet": sheet_name,
                "Table": output_name,
                "Rows": len(frame),
                "Columns": len(frame.columns),
                **audit.get(sheet_name, {"Formula cells": 0, "Populated cells": 0, "Formula fidelity": "Not audited"}),
            }
        )

    pd.DataFrame(sheet_index).to_csv(DATA_DIR / "workbook_sheet_index.csv", index=False)
    print(f"Imported {len(sheet_index)} raw workbook sheets -> data/workbook_*.csv")

    for output_name, (sheet_name, header_marker) in SHEETS.items():
        frame = read_sheet_table(sheet_name, header_marker)
        frame.to_csv(DATA_DIR / f"{output_name}.csv", index=False)
        print(f"Imported {sheet_name} -> data/{output_name}.csv ({len(frame)} rows)")


if __name__ == "__main__":
    main()
